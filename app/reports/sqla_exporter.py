#! ../../venv/bin/python
# -*- coding: utf-8 -*-

import os, sys
from reportlab.platypus import BaseDocTemplate
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Frame, Spacer
from reportlab.platypus import Paragraph as P
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.rl_config import defaultPageSize
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.pagesizes import A0, A1, A2, A3, A4, A5, A6
from reportlab.lib.enums import TA_RIGHT, TA_LEFT, TA_CENTER, TA_JUSTIFY

# set font
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
current_dir = os.path.dirname(os.path.abspath(__file__))
pdfmetrics.registerFont(TTFont('THSarabun', '%s/pdffonts/THSarabun.ttf' % current_dir))
pdfmetrics.registerFont(TTFont('THSarabunBd', '%s/pdffonts/THSarabun Bold.ttf' % current_dir))
pdfmetrics.registerFont(TTFont('THSarabunIt', '%s/pdffonts/THSarabun Italic.ttf' % current_dir))
pdfmetrics.registerFont(TTFont('THSarabunBI', '%s/pdffonts/THSarabun BoldItalic.ttf' % current_dir))
pdfmetrics.registerFontFamily('THSarabun', normal='THSarabun',
        bold='THSarabunBd', italic='THSarabunIt', boldItalic='THSarabunBI')

from sqlalchemy.sql import sqltypes
from copy import copy
import tempfile
import StringIO as io

# helper
def __conv_col_name_to_display(col_name):
    return col_name.title().replace('_', ' ')

def __build_meta_cols(model):
    display_columns = model._display_columns
    __TYPE_CONST = ('datetime', 'date', 'decimal', 'integer', 'boolean', 'string')
    meta_col = []
    for i, disp_conf in enumerate(display_columns):
        col_name, conf = None, {}
        if isinstance(disp_conf, tuple):
            col_name, conf = disp_conf
        else:
            col_name = disp_conf

        col_def = getattr(model.__mapper__.c, col_name, None)
        assert col_def is not None
        display_name = conf.get('text')
        if not bool(display_name): display_name = __conv_col_name_to_display(col_name) # use default

        # simplify type
        if conf.get('type'): t = conf.get('type')
        else:
            _t = col_def.type
            if isinstance(_t, sqltypes.DateTime):
                t = 'datetime'
            elif isinstance(_t, sqltypes.Date):
                t = 'date'
            elif isinstance(_t, sqltypes.Numeric):
                t = 'decimal'
            elif isinstance(_t, sqltypes.Integer):
                t = 'integer'
            elif isinstance(_t, sqltypes.Boolean):
                t = 'boolean'
            else:
                t = 'string'

        assert t in __TYPE_CONST

        # cached config for used below
        # meta schema
        _meta = { 'name': col_name, 'display_name': display_name, 
                  'config': conf, 'sqla_mapper': col_def, 'type': t }
        meta_col.append(_meta)
    return meta_col

def __build_export_cols(meta_col):
    return [e for e in meta_col if not bool(e.get('config').get('not_print'))]

# gen pdf using sqlalchemy adapter
def gen_pdf(res, **kwargs):
    inq_params = kwargs
    m = kwargs.get('model')
    display_columns = m._display_columns

    # setup font
    font_size = 12
    font_name = 'THSarabunBd'

    Story = []
    pstyle = ParagraphStyle(name='pstyle')
    pstyle.fontName = font_name
    pstyle.fontSize = font_size
    number_style = copy(pstyle)
    number_style.alignment = TA_RIGHT
    header_style = copy(pstyle)
    header_style.alignment = TA_CENTER
    font_size_in_table = font_size # magic number!
    tbstyle = [ ('GRID', (0, 0), (-1,-1), 0.1, (0, 0, 0)),
                ('BACKGROUND', (0, 0), (-1, 0), (0.8, 0.8, 0.8)),
                ('FONT', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), font_size_in_table),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE')]

    # guess table style from sample data row
    sample_data = res[0]
    header_cols = []
    meta_col = __build_meta_cols(m)

    # columns that willing to be printed
    export_cols = __build_export_cols(meta_col)
    for i, _meta in enumerate(export_cols):
        header_text = _meta.get('display_name')
        header_cols.append(P("<b>%s</b>" % header_text, header_style))

        v = getattr(sample_data, _meta.get('name'), '')
        _col_type = _meta.get('type')
        if _col_type in ('decimal', 'integer'):
            s = 'ALIGN', (i, 1), (i, -1), 'RIGHT'
            tbstyle.append(s)
        elif _col_type == 'boolean':
            s = 'ALIGN', (i, 1), (i, -1), 'CENTER'
            tbstyle.append(s)

    data = [header_cols]
    # init columns width list having length as list of display columns
    # as the minimum width of word in header text
    col_w = [h.minWidth() for h in header_cols]

    g = {} # global
    def build_rows(rs, options={}):
        __rows = []
        for n, c in enumerate(rs):
            if n % 2 == 1:
                at_row = len(data) + len(__rows)
                tbstyle.append(('BACKGROUND', (0, at_row), (-1, at_row), (0.95, 0.95, 0.95)))

            row = []
            for i, _meta in enumerate(export_cols):
                v = getattr(c, _meta.get('name'), '')
                _col_type = _meta.get('type')
                if _col_type == 'decimal':
                    if v: v = '{:,.2f}'.format(v)
                    else: v = '-'
                elif _col_type == 'integer': v = '{:,.0f}'.format(v)
                elif _col_type == 'boolean': v = ('no', 'false')[v]
                else:
                    v = v if bool(v) else "-"
                    v = v.encode('utf-8') if type(v) is unicode else str(v)
                row.append(v)
            __rows.append(row)

            # calc proper with for every cell
            for i, v in enumerate(col_w):
                w = pdfmetrics.stringWidth(row[i], font_name, font_size)
                if col_w[i] < w: col_w[i] = w
        return __rows

    if bool(m._group_field):
        _res = dict()
        # group records by group field
        for _r in res:
            if getattr(_r, m._group_field) in _res: _res[getattr(_r, m._group_field)].append(_r)
            else: _res[getattr(_r, m._group_field)] = [_r]

        aggr_cols = [ c for c in meta_col if type(c.get('config')) is dict and c.get('config').get('summaryType', None) ]
        print aggr_cols
        # print aggr_cols
        for _g, rs in _res.iteritems():
            at_row = len(data)
            # span colomns at grouping row
            tbstyle.append(('SPAN', (0, at_row), (-1, at_row)))
            tbstyle.append(('BACKGROUND', (0, at_row), (-1, at_row), (0.85, 0.75, 0.65))) #TODO
            tbstyle.append(('FACE', (0, at_row), (-1, at_row), 'THSarabunBI'))

            # data for group header template
            # for providing more flexible on setting template, column name and count of record set
            _d = dict()
            _d[m._group_field] = _g
            _d['c'] = len(rs)

            temp = getattr(m, '_group_header_template')
            if not bool(temp): temp =  "Group By {%s} with {c} record(s)" % m._group_field
            data.append([temp.format(**_d)])

            rows = build_rows(rs)
            data = data + rows
            at_row = len(data)

            # grouping row building
            group_text = []
            for i, _meta in enumerate(aggr_cols):
                _config = _meta.get('config')
                _c = _meta.get('name')
                _fn = _config.get('summaryType').lower()
                _stxt = _config.get('summaryText')
                if not bool(_stxt): _stxt = "%s %s: {value}" % (_fn.title(), _meta.get('display_name'))
                tbstyle.append(('SPAN', (0, at_row), (-1, at_row)))
                tbstyle.append(('ALIGN', (0, at_row), (-1, at_row), 'RIGHT'))
                tbstyle.append(('BACKGROUND', (0, at_row), (-1, at_row), (0.65, 0.75, 0.85)))

                # list of none value of summary column
                _l = [_v for _v in [getattr(v, _c) for v in rs] if bool(_v)]
                # prepare data for binding in template
                _d = dict()
                if _fn == 'sum':
                    # row span
                    _d['raw_value'] = sum(_l)
                    _d['value'] = '{:,.2f}'.format(sum(_l))
                elif _fn == 'average' or _fn == 'avg':
                    if len(_l) == 0:
                        _d['raw_value'] = 0
                        _d['value'] = 0
                    else:
                        _d['raw_value'] = float(sum(_l)) / float(len(_l))
                        _d['value'] = '{:,.2f}'.format(_d['raw_value'])
                elif _fn == 'min': 
                    if len(_l) == 0:
                        _d['raw_value'] = 0
                        _d['value'] = 0
                    else:
                        _d['raw_value'] = min(_l)
                        _d['value'] = '{:,.2f}'.format(_d['raw_value'])
                elif _fn == 'max': 
                    if len(_l) == 0:
                        _d['raw_value'] = 0
                        _d['value'] = 0
                    else:
                        _d['raw_value'] = max(_l)
                        _d['value'] = '{:,.2f}'.format(_d['raw_value'])
                elif _fn == 'count': _d['value'] = len(rs)
                else: raise "Not found aggregate function"
                group_text.append(_stxt.format(**_d))
            print group_text
            if len(group_text) > 0: data.append([(" "*10).join(group_text)])
    else:
        # not group field
        rows = build_rows(rs)
        data = data + rows

    # magic offset !!
    col_w = [v + (pstyle.fontSize + 2) for v in col_w]
    rh = len(data) * [pstyle.fontSize*2 + 5]
    rh = len(data) * [None]
    tbstyle = TableStyle(tbstyle)
    t = Table(data, colWidths=col_w, style=tbstyle, repeatRows=1)
    Story.append(t)

    # prepare print parameters
    pdf_size = eval(getattr(m,'_pdf_size', 'A4'))
    pdf_layout = getattr(m, '_pdf_layout', 'p') # default portrait
    if pdf_layout == 'l':
        pdf_size = (pdf_size[1], pdf_size[0])

    PAGE_WIDTH = pdf_size[0]
    PAGE_HEIGHT = pdf_size[1]

    # event callback
    def draw_header_footer(canv, doc):
        model = inq_params.get('model')
        title = model._title
        condition_text = inq_params.get('condition_text', 'No Filter')
        canv.saveState()

        canv.setFont(font_name, 14)
        canv.drawCentredString(PAGE_WIDTH/2.0, PAGE_HEIGHT - 40, title)

        canv.setFont(font_name, 12)
        canv.drawCentredString(PAGE_WIDTH/2.0, PAGE_HEIGHT - 60, "Filter By %s" % condition_text)

        canv.drawString(PAGE_WIDTH - 40, PAGE_HEIGHT - 25, "Page %s" % doc.page)
        canv.restoreState()
        # canv.setFont(font_name, 12)
    
    out = io.StringIO()
    doc = SimpleDocTemplate(out, pagesize=pdf_size, topMargin=65, bottomMargin=10, leftMargin=10, rightMargin=10)
    doc.build(Story, onFirstPage=draw_header_footer, onLaterPages=draw_header_footer)
    pdf_stream = out.getvalue()
    out.seek(0)
    # out.close()
    return out

from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.cell import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Style
def gen_xl(res, **kwargs):
    m = kwargs.get("model")
    meta_col = __build_meta_cols(m)
    export_cols = __build_export_cols(meta_col)

    wb = Workbook()
    ws = wb.active # worksheet
    ws.title = "Excel Using Openpyxl"
    row = 1 # start row
    col = 1 # col
    
    # prepared formatters
    def __number_cell_format(cell):
        cell.style = Style(number_format='0.00')
    def __int_cell_format(cell):
        cell.style = Style(number_format='0')
    def __dummy_cell_format(cell):
        pass

    formatters = []
    # write header row and gather formatters
    for i, _meta in enumerate(export_cols):
        col_idx = i + 1
        ws.cell(row=row, column=col_idx).value = _meta.get('display_name')
        _t = _meta.get('type')
        if _t == 'decimal': formatters.append(__number_cell_format)
        elif _t == 'integer': formatters.append(__int_cell_format)
        else: formatters.append(__dummy_cell_format)

    row += 1
    for i, r in enumerate(res):
        for j, _meta in enumerate(export_cols):
            col_idx = j + 1
            v = getattr(r, _meta.get('name'))
            cell = ws.cell(row=row, column=col_idx)
            cell.value = v
            formatters[j](cell)
        row += 1

    # save workbook to stream
    out = io.StringIO()
    wb.save(out)
    out.seek(0)
    return out

def gen_csv(res, **kwargs):
    pass

# TEST -----------------------------------------------------------
# lib_path = os.path.abspath('../../')
# sys.path.append(lib_path)
# from manage import app, db
# from cmms import create_app
# from cmms.models import db
# from flask.ext.sqlalchemy import SQLAlchemy
# app = create_app("cmms.settings.DevConfig", 'dev')
# db.init_app(app)
# from cmms.models import Client

# with app.app_context():
#     clients = Client.query.filter("total_credit_limit > 0").order_by("total_credit_limit desc").limit(90)
#     go(clients)

